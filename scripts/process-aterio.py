"""
Processes the Aterio US Data Centers Dashboard XLSX into two JSON files
used by the slideshow app. Run this script whenever the source XLSX is updated.

Outputs:
  public/aterio_states.json     — Per-state aggregates (active/planned counts,
                                  power MW, sqft, dominant operator) derived
                                  from the "US Data Centers Dataset" sheet.

  public/aterio_yearly_mw.json  — National cumulative data center power capacity
                                  (MW, baseline estimate) by year 2018–2035,
                                  derived from the "US Additional Info - Evo ISO"
                                  sheet. Used by the slide 2 dual-axis chart.

Dependencies:
  pip install openpyxl   (or use a venv — do NOT install globally)

Run from project root:
  python3 -m venv .venv && .venv/bin/pip install openpyxl -q
  .venv/bin/python3 scripts/process-aterio.py
  rm -rf .venv
"""

import json, sys
from pathlib import Path
from collections import defaultdict, Counter

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl not installed. Run: pip3 install --user openpyxl")

# FIPS lookup (state_code -> id used by us-states.json GeoJSON)
FIPS = {
    'AL':'01','AK':'02','AZ':'04','AR':'05','CA':'06','CO':'08','CT':'09',
    'DE':'10','DC':'11','FL':'12','GA':'13','HI':'15','ID':'16','IL':'17',
    'IN':'18','IA':'19','KS':'20','KY':'21','LA':'22','ME':'23','MD':'24',
    'MA':'25','MI':'26','MN':'27','MS':'28','MO':'29','MT':'30','NE':'31',
    'NV':'32','NH':'33','NJ':'34','NM':'35','NY':'36','NC':'37','ND':'38',
    'OH':'39','OK':'40','OR':'41','PA':'42','RI':'44','SC':'45','SD':'46',
    'TN':'47','TX':'48','UT':'49','VT':'50','VA':'51','WA':'53','WV':'54',
    'WI':'55','WY':'56',
}

PLANNED_STAGES = {'Construction', 'Announcement', 'Delayed'}

wb   = openpyxl.load_workbook('public/Aterio US Data Centers Dashboard.xlsx', read_only=True, data_only=True)
ws   = wb['US Data Centers Dataset']
rows = list(ws.iter_rows(min_row=2, values_only=True))

# Column indices (0-based) from header row 1:
# ATERIO_DATACENTER_UNIQUE_ID, DATA_CENTER_BUILDING_NAME, DATA_CENTER_SCATTER_PLOT_NAME,
# PROVIDER_NAME(3), DATACENTER_STAGE(4), STATE_CODE(5), STATE_NAME(6),
# TOT_FACILITY_SPACE_SQFT(7), TOT_DATACENTER_SPACE_SQFT(8), TOT_POWER_MW(9)

by_state = defaultdict(lambda: {
    'active': 0, 'planned': 0,
    'active_mw': 0.0, 'planned_mw': 0.0,
    'active_sqft': 0, 'operators': Counter(),
    'state_name': '', 'state_code': ''
})

for r in rows:
    provider  = r[3] or ''
    stage     = r[4] or ''
    s_code    = r[5] or ''
    s_name    = r[6] or ''
    sqft      = r[7] or 0
    mw        = r[9] or 0.0

    if not s_code or s_code not in FIPS:
        continue

    d = by_state[s_code]
    d['state_code'] = s_code
    d['state_name'] = s_name

    if stage == 'Active':
        d['active']      += 1
        d['active_mw']   += float(mw)
        d['active_sqft'] += int(sqft) if sqft else 0
        if provider:
            d['operators'][provider] += 1
    elif stage in PLANNED_STAGES:
        d['planned']    += 1
        d['planned_mw'] += float(mw)

out = []
for s_code, d in sorted(by_state.items(), key=lambda x: -x[1]['active']):
    top_op = d['operators'].most_common(1)
    op_name  = top_op[0][0] if top_op else None
    op_count = top_op[0][1] if top_op else 0
    out.append({
        'id':                    FIPS[s_code],
        'state_code':            s_code,
        'state_name':            d['state_name'],
        'active':                d['active'],
        'planned':               d['planned'],
        'active_mw':             round(d['active_mw'], 1),
        'planned_mw':            round(d['planned_mw'], 1),
        'active_sqft':           d['active_sqft'],
        'dominant_operator':     op_name,
        'dominant_operator_n':   op_count,
    })

Path('public/aterio_states.json').write_text(json.dumps(out, indent=2))
print(f"Written {len(out)} states to public/aterio_states.json")

# --- Yearly national cumulative MW from Evo ISO sheet ---
ws_evo   = wb['US Additional Info - Evo ISO']
evo_rows = list(ws_evo.iter_rows(min_row=2, values_only=True))
# BAL_AUTH_ABBR(0), BAL_AUTH_NAME(1), ACTIVATION_YEAR(2), OPTIMISTIC(3), BASELINE(4), CONSERVATIVE(5)
by_year_mw = defaultdict(float)
for r in evo_rows:
    year, baseline = r[2], r[4]
    if year and baseline:
        by_year_mw[int(year)] += float(baseline)

mw_out = [{'year': y, 'mw': round(v, 1)} for y, v in sorted(by_year_mw.items())]
Path('public/aterio_yearly_mw.json').write_text(json.dumps(mw_out, indent=2))
print(f"Written {len(mw_out)} years to public/aterio_yearly_mw.json")

# Quick sanity check
total_active  = sum(d['active']  for d in out)
total_planned = sum(d['planned'] for d in out)
total_mw      = sum(d['active_mw'] for d in out)
print(f"Total active:  {total_active}")
print(f"Total planned: {total_planned}")
print(f"Total MW:      {total_mw:,.0f}")
top5 = sorted(out, key=lambda x: -x['active'])[:5]
print("Top 5 states:", [(d['state_name'], d['active']) for d in top5])
